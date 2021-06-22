import pprint
import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from nornir import InitNornir
from nornir.core.task import Result, Task
from nornir_netmiko.tasks import netmiko_send_command
from nornir_salt.plugins.functions import ResultSerializer

from parsing import ExosParse


# build hosts dictionary from excel file...
def get_hosts_file(excel_file):
    sheet = 'hosts'
    book = load_workbook(excel_file)
    sheet = book[sheet]

    hosts = dict()
    for row in sheet.rows:
        key = row[0].value
        if key:
            hosts[key] = {
                'hostname': row[1].value.strip(),
                'username': row[2].value.strip(),
                'password': row[3].value.strip(),
                'platform': row[4].value.strip(),
                'port': row[5].value,
                'data': {
                    'vendors': row[6].value.strip(),
                    'role': row[7].value.strip(),
                    'site': row[8].value.strip(),
                },
            }

    if hosts['id']:
        del hosts['id']
    inventory_hosts = {'hosts': hosts}

    # pprint.pprint(inventory_hosts)
    return inventory_hosts


# return inventory hostname(in this case, ipaddress)
def device_ip(task: Task) -> Result:
    return Result(
        host=task.host,
        result=f'{task.host.hostname}'
    )


def netmiko_send_commands(task, commands, **kwargs):
    for command in commands:
        task.run(
            task=netmiko_send_command,
            command_string=command,
            name=command,
            **kwargs.get("netmiko_kwargs", {})
        )
    task.run(task=device_ip)
    return Result(host=task.host)


def exos_task(hosts, **kwargs):
    nr = InitNornir(
        runner={
            'plugin': 'threaded',
            'options': {
                'num_workers': 30,
            },
        },
        inventory={
            'plugin': 'DictInventory',
            'options': {
                'hosts': hosts['hosts'],
                'groups': {},
                'defaults': {}
            }
        }
    )

    vendors = kwargs['vendors']
    commands = kwargs['commands']
    site = ''
    role = ''
    if kwargs['site']:
        site = kwargs['site']
    if kwargs['role']:
        role = kwargs['role']

    if not commands:
        print('Need to commands...')

    if site and role:
        nr = nr.filter(vendors='extreme', site=site, role=role)
    else:
        if site:
            nr = nr.filter(vendors='extreme', site=site)
        elif role:
            nr = nr.filter(vendors=vendors, role=role)
        else:
            nr = nr.filter(vendors='extreme')

    # conn_timeout, global_delay_factory didn't work.
    # so i was modify that value into the netmiko base_connection.py
    # each default value was 5, 1
    result = nr.run(task=netmiko_send_commands, commands=commands, conn_timeout=30, global_delay_factor=2)

    return ResultSerializer(result)


# result is serialized Result(dict type).
def data_parsing(result):
    check_list = list()
    # made dictionary to string data.
    for i in result.keys():
        ipa = result[i]['device_ip']
        tmp = dict()
        tmp_str = ''
        for j in result[i].keys():
            if result[i][j]:
                # print(result[i][j].values())
                tmp_str = tmp_str + result[i][j] + '\n'
        exos_parse = ExosParse(tmp_str)
        tmp = {
            'ip': ipa,
            'hostname': exos_parse.hostname(),
            'dev_model': exos_parse.dev_model(),
            'os_version': exos_parse.os_ver(),
            'uptime': exos_parse.uptime(),
            'cpu_idle': exos_parse.cpu_usage(),
            'mem_free': exos_parse.mem_usage(),
            'fan': exos_parse.fan(),
            'temperature': exos_parse.temperature(),
            'power': exos_parse.power_supply(),
        }
        check_list.append(tmp)
    return check_list


def create_worksheet(excel_file):
    today = datetime.date.today()
    book = load_workbook(excel_file)
    sheet_title = 'report_%s' % (today,)
    header = ['IP', 'Host Name', 'Model', 'Ver.', 'UP Time', 'CPU(% idle)', 'MEM(% free)', 'Fan', 'Temp.', 'Power']
    try:
        book.create_sheet(title=sheet_title)
        sheet = book[sheet_title]
        # sheet.merge_cells('A1:J1')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        sheet['A1'] = 'Proactive Maintenance Report - %s' % (today,)
        sheet['A1'].font = Font(size=14, bold=True, underline='single')
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.append(header)
        book.save(excel_file)
        book.close()
    except Exception as e:
        print('Some error occurred. \n  %s' % (e,))
        book.close()
    finally:
        book.close()


def report(check_list, excel_file):
    today = datetime.date.today()
    book = load_workbook(excel_file)
    sheet_title = f'report_{today}'
    sheet = book[sheet_title]
    for i in check_list:
        data = [
            i['ip'],
            i['hostname'],
            i['dev_model'],
            i['os_version'],
            i['uptime'],
            i['cpu_idle'],
            i['mem_free'],
            i['fan'],
            i['temperature'],
            i['power'],
        ]
        sheet.append(data)
    book.save(excel_file)
    book.close()


if __name__ == '__main__':
    now = datetime.datetime.now()
    print('>>>>>>>>>>>>>>>>>>>>>>>>')
    print(f'Current time: {now}')

    commands = [
        'show switch', 'show fan', 'show temp', 'show power', 'show cpu',
        'show memory', 'show config | inc sysName'
    ]
    hosts = get_hosts_file('exos_pm.xlsx')
    result = exos_task(hosts, commands=commands, vendors='extreme', site='wifi', role='l3')
    pprint.pprint(result)
    chk_list = data_parsing(result)
    pprint.pprint(chk_list)

    print('<<<<<<<<<<<<<<<<<<<<<<<<<')
    now = datetime.datetime.now()
    print(f'Current time: {now}')
